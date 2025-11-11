#!/usr/bin/env python3
"""Vergleich von Nachunternehmer-Angeboten.

Dieses Skript liest mehrere Angebotsdateien im CSV- oder Excel-Format ein,
ordnet Positionen anhand definierter Schlüsselspalten zu und erzeugt eine
Excel-Auswertung mit farblichen Hervorhebungen und einer Gesamtsumme pro
Anbieter.
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class OfferData:
    provider: str
    price_column: str
    price_df: pd.DataFrame
    key_df: pd.DataFrame


NORMALIZED_KEY_CANDIDATES = {
    "lv": [
        "lv",
        "lv_nr",
        "lvnr",
        "lvnummer",
        "lv_nummer",
        "lv-nr",
        "position",
        "pos",
        "pos_nr",
        "posnr",
        "positionsnummer",
        "item",
        "item_no",
        "itemnr",
    ],
    "description": [
        "bezeichnung",
        "beschreibung",
        "leistung",
        "positionstext",
        "text",
        "kurztext",
        "titel",
    ],
}

PRICE_CANDIDATES = [
    "gesamtpreis",
    "gesamt",
    "positionspreis",
    "summe",
    "preis",
    "betrag",
    "total",
    "totalpreis",
    "amount",
]

QUANTITY_CANDIDATES = [
    "menge",
    "qty",
    "quantity",
    "anzahl",
]

UNIT_PRICE_CANDIDATES = [
    "einheitspreis",
    "ep",
    "einzelpreis",
    "preis_e",
    "preis_pro_einheit",
    "unit_price",
    "preis_pro_stueck",
]


def normalize_column_name(name: str) -> str:
    cleaned = "".join(ch.lower() if ch.isalnum() else "_" for ch in name)
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_")


def read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xls", ".xlsx", ".xlsm"}:
        return pd.read_excel(path)
    if path.suffix.lower() in {".csv", ".txt"}:
        return pd.read_csv(path, sep=None, engine="python")
    raise ValueError(f"Unsupported file format: {path.suffix}")


def unique_provider_name(existing: Iterable[str], candidate: str) -> str:
    base = candidate
    counter = 1
    while candidate in existing:
        counter += 1
        candidate = f"{base}_{counter}"
    return candidate


def find_matching_column(
    normalized_to_original: Dict[str, str],
    candidates: Sequence[str],
) -> Optional[str]:
    for candidate in candidates:
        if candidate in normalized_to_original:
            return normalized_to_original[candidate]
    return None


def resolve_key_columns(
    normalized_to_original: Dict[str, str],
    requested_keys: Optional[Sequence[str]],
) -> List[Tuple[str, str]]:
    key_columns: List[Tuple[str, str]] = []

    if requested_keys:
        for key in requested_keys:
            normalized_key = normalize_column_name(key)
            match = find_matching_column(normalized_to_original, [normalized_key])
            if not match:
                raise KeyError(
                    f"Angeforderte Schlüsselspalte '{key}' wurde in den Daten nicht gefunden."
                )
            key_columns.append((normalized_key, match))
        return key_columns

    for normalized_key, candidates in NORMALIZED_KEY_CANDIDATES.items():
        match = find_matching_column(normalized_to_original, candidates)
        if match:
            key_columns.append((normalized_key, match))

    if not key_columns:
        raise KeyError(
            "Es konnten keine passenden Schlüsselspalten gefunden werden. "
            "Bitte mit --keys explizit angeben."
        )

    return key_columns


def build_offer_data(
    path: Path,
    requested_keys: Optional[Sequence[str]],
    display_name_map: Dict[str, str],
) -> OfferData:
    table = read_table(path)
    if table.empty:
        raise ValueError(f"Datei '{path}' enthält keine Daten.")

    normalized_to_original = {
        normalize_column_name(col): col for col in table.columns
    }

    key_columns = resolve_key_columns(normalized_to_original, requested_keys)

    price_col = find_matching_column(normalized_to_original, PRICE_CANDIDATES)

    if not price_col:
        qty_col = find_matching_column(normalized_to_original, QUANTITY_CANDIDATES)
        unit_col = find_matching_column(normalized_to_original, UNIT_PRICE_CANDIDATES)
        if qty_col and unit_col:
            qty = pd.to_numeric(table[qty_col], errors="coerce")
            unit_price = pd.to_numeric(table[unit_col], errors="coerce")
            table["__calculated_total__"] = qty * unit_price
            price_col = "__calculated_total__"
        else:
            raise KeyError(
                f"In Datei '{path}' wurde keine Preis-Spalte gefunden. "
                "Bitte stellen Sie sicher, dass eine Summe, Menge und Einheitspreis vorhanden sind."
            )

    price_series = pd.to_numeric(table[price_col], errors="coerce")

    key_df = pd.DataFrame(index=table.index)
    key_parts: List[pd.Series] = []

    for normalized_key, original in key_columns:
        display_name = display_name_map.setdefault(normalized_key, original)
        series = table[original]
        series = series.where(series.notna(), "")
        series = series.astype(str).str.strip()
        key_df[display_name] = series
        key_parts.append(series)

    for normalized_key, display_name in list(display_name_map.items()):
        if normalized_key == "__providers__":
            continue
        if display_name not in key_df.columns:
            key_df[display_name] = ""

    if not key_parts:
        raise ValueError("Keine Schlüsselspalten verfügbar.")

    combined_key = key_parts[0]
    for part in key_parts[1:]:
        combined_key = combined_key + " | " + part

    combined_key = combined_key.str.replace(r"\bnan\b", "", regex=True)
    key_df["__key__"] = combined_key.str.strip(" |")

    valid_mask = key_df["__key__"].astype(bool) & price_series.notna()
    key_df = key_df.loc[valid_mask].copy()
    price_series = price_series.loc[valid_mask]

    price_df = pd.DataFrame({"__key__": key_df["__key__"], "price": price_series})

    provider_name = unique_provider_name(
        existing=display_name_map.get("__providers__", []),
        candidate=os.path.splitext(path.name)[0],
    )

    providers = display_name_map.setdefault("__providers__", [])
    providers.append(provider_name)

    return OfferData(
        provider=provider_name,
        price_column=f"{provider_name} Preis",
        price_df=price_df,
        key_df=key_df,
    )


def merge_offers(offers: List[OfferData], key_column_names: List[str]) -> pd.DataFrame:
    if not offers:
        raise ValueError("Es wurden keine Angebote geladen.")

    key_frames = []
    for offer in offers:
        key_frame = offer.key_df[[*key_column_names, "__key__"]].copy()
        key_frames.append(key_frame)

    base = (
        pd.concat(key_frames, ignore_index=True)
        .drop_duplicates(subset="__key__")
        .reset_index(drop=True)
    )

    result = base
    for offer in offers:
        price_df = offer.price_df.rename(columns={"price": offer.price_column})
        result = result.merge(price_df, on="__key__", how="left")

    return result


def format_excel(
    writer: pd.ExcelWriter,
    data_df: pd.DataFrame,
    price_columns: List[str],
    diff_columns_eur: List[str],
    diff_columns_pct: List[str],
    summary_df: pd.DataFrame,
) -> None:
    workbook = writer.book

    data_sheet = writer.sheets["Positionen"]
    summary_sheet = writer.sheets["Übersicht"]

    currency_format = "#,##0.00"
    percent_format = "0.00%"

    max_row = data_df.shape[0] + 1

    # Apply number formats
    for col_name in price_columns + diff_columns_eur:
        if col_name not in data_df.columns:
            continue
        col_idx = data_df.columns.get_loc(col_name) + 1
        column_letter = get_column_letter(col_idx)
        for row in range(2, max_row + 1):
            cell = data_sheet[f"{column_letter}{row}"]
            cell.number_format = currency_format

    for col_name in diff_columns_pct:
        if col_name not in data_df.columns:
            continue
        col_idx = data_df.columns.get_loc(col_name) + 1
        column_letter = get_column_letter(col_idx)
        for row in range(2, max_row + 1):
            cell = data_sheet[f"{column_letter}{row}"]
            cell.number_format = percent_format

    # Highlight cheapest provider in price columns
    if price_columns:
        price_letters = [
            get_column_letter(data_df.columns.get_loc(col) + 1) for col in price_columns
        ]
        min_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for letter in price_letters:
            cell_range = f"{letter}2:{letter}{max_row}"
            formula = f"{letter}2=MIN(${price_letters[0]}2:${price_letters[-1]}2)"
            data_sheet.conditional_formatting.add(
                cell_range,
                FormulaRule(formula=[formula], fill=min_fill),
            )

    # Conditional formatting for diff columns
    positive_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    negative_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col_list in (diff_columns_eur, diff_columns_pct):
        for col_name in col_list:
            if col_name not in data_df.columns:
                continue
            col_idx = data_df.columns.get_loc(col_name) + 1
            letter = get_column_letter(col_idx)
            cell_range = f"{letter}2:{letter}{max_row}"
            data_sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="lessThan", formula=["0"], fill=negative_fill),
            )
            data_sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="greaterThan", formula=["0"], fill=positive_fill),
            )

    # Autofit columns (approximation by adjusting width based on max length)
    for column_cells in data_sheet.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        if not values:
            continue
        max_length = max(len(str(value)) for value in values)
        adjusted_width = min(60, max_length + 2)
        data_sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Summary sheet formatting
    summary_max_row = summary_df.shape[0] + 1
    summary_price_col = get_column_letter(summary_df.columns.get_loc("Gesamtpreis") + 1)

    for row in range(2, summary_max_row + 1):
        cell = summary_sheet[f"{summary_price_col}{row}"]
        cell.number_format = currency_format

    diff_eur_col = summary_df.columns.get_loc("Diff €") + 1
    diff_pct_col = summary_df.columns.get_loc("Diff %") + 1

    for row in range(2, summary_max_row + 1):
        summary_sheet[f"{get_column_letter(diff_eur_col)}{row}"].number_format = currency_format
        summary_sheet[f"{get_column_letter(diff_pct_col)}{row}"].number_format = percent_format

    if not summary_df.empty:
        min_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        summary_sheet.conditional_formatting.add(
            f"{summary_price_col}2:{summary_price_col}{summary_max_row}",
            FormulaRule(
                formula=[
                    f"{summary_price_col}2=MIN(${summary_price_col}$2:${summary_price_col}${summary_max_row})"
                ],
                fill=min_fill,
            ),
        )

    header_font = Font(bold=True)
    for sheet in (data_sheet, summary_sheet):
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Vergleicht mehrere Nachunternehmer-Angebote und erzeugt eine Excel-Auswertung.",
    )
    parser.add_argument(
        "output",
        type=Path,
        help="Pfad zur Ergebnis-Excel-Datei",
    )
    parser.add_argument(
        "offers",
        nargs="+",
        type=Path,
        help="Pfad zu Angebotsdateien (CSV oder Excel)",
    )
    parser.add_argument(
        "--keys",
        nargs="*",
        help="Optional: Namen der Spalten, die zur Zuordnung verwendet werden sollen.",
    )

    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)

    display_name_map: Dict[str, str] = {}
    offers: List[OfferData] = []

    for offer_path in args.offers:
        if not offer_path.exists():
            raise FileNotFoundError(f"Datei '{offer_path}' wurde nicht gefunden.")
        offer = build_offer_data(offer_path, args.keys, display_name_map)
        offers.append(offer)

    provider_names = [offer.provider for offer in offers]

    key_column_names = [name for key, name in display_name_map.items() if key not in {"__providers__"}]

    result_df = merge_offers(offers, key_column_names)

    price_columns = [offer.price_column for offer in offers]

    # Berechnungen für günstigsten Anbieter und Differenzen
    min_price_col = "Günstigster Preis"
    if price_columns:
        result_df[min_price_col] = result_df[price_columns].min(axis=1, skipna=True)

        def best_provider(row: pd.Series) -> Optional[str]:
            min_price = row[min_price_col]
            if pd.isna(min_price):
                return None
            for provider, col in zip(provider_names, price_columns):
                value = row[col]
                if pd.notna(value) and abs(value - min_price) <= 1e-6:
                    return provider
            return None

        result_df["Günstigster Anbieter"] = result_df.apply(best_provider, axis=1)

        diff_columns_eur: List[str] = []
        diff_columns_pct: List[str] = []

        for provider, col in zip(provider_names, price_columns):
            diff_eur_col = f"{provider} Diff €"
            diff_pct_col = f"{provider} Diff %"
            diff = result_df[col] - result_df[min_price_col]
            result_df[diff_eur_col] = diff
            result_df[diff_pct_col] = np.where(
                result_df[min_price_col].abs() > 1e-9,
                diff / result_df[min_price_col],
                np.nan,
            )
            diff_columns_eur.append(diff_eur_col)
            diff_columns_pct.append(diff_pct_col)
    else:
        diff_columns_eur = []
        diff_columns_pct = []

    # Zusammenfassung
    summary_rows = []
    totals = []
    for provider, col in zip(provider_names, price_columns):
        total = result_df[col].sum(skipna=True)
        totals.append(total)
        summary_rows.append({
            "Anbieter": provider,
            "Gesamtpreis": total,
        })

    if summary_rows:
        valid_totals = [total for total in totals if pd.notna(total)]
        min_total = min(valid_totals) if valid_totals else None
        for row in summary_rows:
            total = row["Gesamtpreis"]
            if min_total is None or pd.isna(total):
                row["Diff €"] = np.nan
                row["Diff %"] = np.nan
                continue
            diff = total - min_total
            row["Diff €"] = diff
            row["Diff %"] = diff / min_total if abs(min_total) > 1e-9 else np.nan
    summary_df = pd.DataFrame(summary_rows)

    # Ausgabereihenfolge festlegen
    column_order = [name for name in key_column_names if name in result_df.columns]
    column_order += price_columns
    if min_price_col in result_df.columns:
        column_order.append(min_price_col)
    if "Günstigster Anbieter" in result_df.columns:
        column_order.append("Günstigster Anbieter")
    column_order += diff_columns_eur + diff_columns_pct
    result_df = result_df[column_order]

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="Positionen", index=False)
        summary_df.to_excel(writer, sheet_name="Übersicht", index=False)
        format_excel(
            writer=writer,
            data_df=result_df,
            price_columns=price_columns,
            diff_columns_eur=diff_columns_eur,
            diff_columns_pct=diff_columns_pct,
            summary_df=summary_df,
        )

    print(f"Auswertung gespeichert unter: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
